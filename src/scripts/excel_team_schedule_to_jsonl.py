#!/usr/bin/env python3
"""
Convert Throw Down team sheets into per-team games.jsonl files.

Each team sheet lists PLAYING assignments (HOME, AWAY, or stream court).
Unlike the court converter, this includes every game a team plays — including
Court 1 stream matchups and AWAY assignments.

Output files: {date}_{team_slug}.jsonl in the output directory.

Usage:
  python excel_team_schedule_to_jsonl.py schedule/master_schedule.xlsx \\
    --output-dir src/output/June2026Tournament/schedule/generated_teams
"""

import argparse
import json
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from excel_schedule_to_jsonl import (
    THROWDOWN_SKIP_SHEETS,
    build_team_aliases,
    normalize_time,
    resolve_team_name,
)
from throwdown_bracket import build_bracket_time_map, classify_game


PLAYING_RE = re.compile(
    r"PLAYING\s*:\s*Court\s*(\d+)\s*(?:\([^)]*\))?\s*vs\.?\s*(.+)",
    re.I,
)


def team_slug(name):
    return re.sub(r"[^a-z0-9]+", "_", str(name).lower()).strip("_")


def playing_role(assignment):
    upper = assignment.upper()
    if "(HOME)" in upper or "STREAM COURT - HOME" in upper:
        return "home"
    if "(AWAY)" in upper or "STREAM COURT - AWAY" in upper:
        return "away"
    return "home"


def list_team_sheets(workbook, teams_filter=None):
    sheets = []
    for sheet_name in workbook.sheetnames:
        if sheet_name in THROWDOWN_SKIP_SHEETS:
            continue
        if teams_filter and team_slug(sheet_name) not in teams_filter:
            slug_names = {team_slug(t) for t in teams_filter}
            if team_slug(sheet_name) not in slug_names:
                # Also allow matching by display name
                if sheet_name not in teams_filter:
                    continue
        sheets.append(sheet_name)
    return sheets


def parse_team_sheet(workbook, sheet_name, date, minutes, aliases, bracket_start, bracket_time_map):
    games = []
    seen = set()

    for row in workbook[sheet_name].iter_rows(values_only=True):
        if not row or len(row) < 3:
            continue
        assignment = row[2]
        if not assignment or not isinstance(assignment, str):
            continue
        if "PLAYING" not in assignment.upper():
            continue

        match = PLAYING_RE.search(assignment)
        if not match:
            continue

        court_num = int(match.group(1))
        opponent = resolve_team_name(match.group(2).strip(), aliases)
        role = playing_role(assignment)

        round_val = row[0]
        time_val = row[1]
        if round_val is None or time_val is None:
            continue

        start_time = normalize_time(time_val)
        if not start_time:
            continue

        round_num = int(round_val)
        game_type, round_label = classify_game(
            start_time,
            round_num,
            bracket_start,
            bracket_time_map,
            bracket_minutes=minutes,
        )
        if role == "home":
            home_team = sheet_name
            away_team = opponent
        else:
            home_team = opponent
            away_team = sheet_name

        dedupe_key = (round_num, start_time, court_num, home_team, away_team)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        games.append({
            "date": date,
            "team": sheet_name,
            "team_slug": team_slug(sheet_name),
            "court": f"court{court_num}",
            "court_display": f"Court {court_num}",
            "type": game_type,
            "round": round_label,
            "home_team": home_team,
            "away_team": away_team,
            "start_time": start_time,
            "minutes": minutes,
        })

    games.sort(key=lambda g: (g["start_time"], g["round"]))
    return games


def write_team_jsonl_files(games, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    grouped = defaultdict(list)
    for game in games:
        key = (game["date"], game["team_slug"])
        grouped[key].append(game)

    written = []
    for (date, slug), team_games in sorted(grouped.items()):
        outfile = output_dir / f"{date}_{slug}.jsonl"
        with open(outfile, "w") as f:
            for game in team_games:
                record = {
                    "type": game.get("type", "round_robin"),
                    "round": game.get("round", ""),
                    "home_team": game["home_team"],
                    "away_team": game["away_team"],
                    "start_time": game["start_time"],
                    "minutes": game["minutes"],
                    "court": game["court_display"],
                    "team": game["team"],
                }
                f.write(json.dumps(record) + "\n")
        written.append(outfile)
        print(f"Wrote {len(team_games)} games to {outfile}")

    return written


def parse_teams_filter(value):
    if not value:
        return None
    return [part.strip() for part in value.split(",") if part.strip()]


def main():
    parser = argparse.ArgumentParser(
        description="Convert Throw Down team sheets to per-team games.jsonl files"
    )
    parser.add_argument("excel_file", help="Path to tournament schedule .xlsx")
    parser.add_argument(
        "--output-dir",
        default="src/output/June2026Tournament/schedule/generated_teams",
        help="Directory for generated team JSONL files",
    )
    parser.add_argument(
        "--date",
        default="2026-06-20",
        help="Tournament date for team sheets (default: 2026-06-20)",
    )
    parser.add_argument(
        "--minutes",
        type=int,
        default=25,
        help="Game duration in minutes (default: 25)",
    )
    parser.add_argument(
        "--teams",
        default=None,
        help="Comma-separated team names or slugs (default: all team sheets)",
    )
    parser.add_argument(
        "--copy-to",
        default=None,
        help="Copy source Excel to this path (e.g. schedule/master_schedule.xlsx)",
    )
    parser.add_argument(
        "--list-teams",
        action="store_true",
        help="Print team sheet names and slugs, then exit",
    )
    parser.add_argument(
        "--bracket-start",
        default=None,
        help="Wall-clock bracket start (default: detect from workbook or date defaults)",
    )
    parser.add_argument(
        "--round-robin-max",
        type=int,
        default=10,
        help="Highest round-robin round number before times imply bracket (default: 10)",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"Error: Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    if args.copy_to:
        copy_dest = Path(args.copy_to)
        copy_dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(excel_path, copy_dest)
        print(f"Copied schedule to {copy_dest}")

    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    teams_filter = parse_teams_filter(args.teams)
    team_sheets = list_team_sheets(workbook, teams_filter)

    if args.list_teams:
        for name in team_sheets:
            print(f"{name}\t{team_slug(name)}")
        return

    if not team_sheets:
        print("Error: No team sheets matched.", file=sys.stderr)
        sys.exit(1)

    aliases = build_team_aliases(workbook)
    bracket_start, bracket_time_map = build_bracket_time_map(workbook, args.date)
    if args.bracket_start:
        bracket_start = normalize_time(args.bracket_start) or args.bracket_start
    print(f"Bracket start: {bracket_start} ({len(bracket_time_map)} known slot(s))")

    all_games = []
    for sheet_name in team_sheets:
        games = parse_team_sheet(
            workbook,
            sheet_name,
            args.date,
            args.minutes,
            aliases,
            bracket_start,
            bracket_time_map,
        )
        all_games.extend(games)
        print(f"Parsed {len(games)} games for {sheet_name}")

    if not all_games:
        print("Error: No games parsed from team sheets.", file=sys.stderr)
        sys.exit(1)

    write_team_jsonl_files(all_games, args.output_dir)


if __name__ == "__main__":
    main()
