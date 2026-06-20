#!/usr/bin/env python3
"""
Convert a tournament Excel schedule into per-court games.jsonl files.

Supports:
  - table format: flat sheet with date, court, home_team, away_team, start_time, minutes
  - throwdown format: BDL Throw Down team sheets with PLAYING assignments

Output files: {date}_court{N}.jsonl in the output directory.

Usage:
  python excel_schedule_to_jsonl.py schedule/master_schedule.xlsx \\
    --output-dir src/output/June2026Tournament/schedule/generated
"""

import argparse
import json
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

from throwdown_bracket import build_bracket_time_map, classify_game, merge_bracket_games

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


THROWDOWN_SKIP_SHEETS = {"Overview Schedule", "BOARD Schedule", "BLANK"}
PLAYING_HOME_RE = re.compile(
    r"PLAYING\s*:\s*Court\s*(\d+)\s*\(HOME\)\s*vs\.?\s*(.+)",
    re.I,
)
PLAYING_NEUTRAL_RE = re.compile(
    r"PLAYING\s*:\s*Court\s*(\d+)\s*(?!\([^)]*\))\s*vs\.?\s*(.+)",
    re.I,
)

COLUMN_ALIASES = {
    "date": "date",
    "day": "date",
    "court": "court",
    "type": "type",
    "game_type": "type",
    "round": "round",
    "home_team": "home_team",
    "home": "home_team",
    "away_team": "away_team",
    "away": "away_team",
    "start_time": "start_time",
    "time": "start_time",
    "minutes": "minutes",
    "duration": "minutes",
    "duration_minutes": "minutes",
}


def normalize_header(value):
    if value is None:
        return None
    key = str(value).strip().lower().replace(" ", "_")
    return COLUMN_ALIASES.get(key)


def normalize_date(value):
    text = str(value).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", text):
        return text
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", text)
    if match:
        month, day, year = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"
    return text


def normalize_court(value):
    text = str(value).strip().lower()
    match = re.search(r"(\d+)", text)
    if match:
        return f"court{match.group(1)}"
    return text.replace(" ", "")


def normalize_time(value):
    if value is None or str(value).strip() == "":
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    text = str(value).strip()
    if re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", text):
        parts = text.split(":")
        if len(parts) == 2:
            return f"{int(parts[0]):02d}:{parts[1]}"
        return f"{int(parts[0]):02d}:{parts[1]}:{parts[2]}"
    return text


def normalize_team_key(name):
    return re.sub(r"[^a-z0-9]+", "", str(name).lower())


def build_team_aliases(workbook):
    aliases = {}
    for sheet_name in workbook.sheetnames:
        if sheet_name in THROWDOWN_SKIP_SHEETS:
            continue
        aliases[normalize_team_key(sheet_name)] = sheet_name
    aliases[normalize_team_key("That's So Raven")] = aliases.get(
        normalize_team_key("Thats So Raven"), "Thats So Raven"
    )
    aliases[normalize_team_key("Cleveland Show")] = aliases.get(
        normalize_team_key("The Cleveland Show"), "The Cleveland Show"
    )
    return aliases


def resolve_team_name(raw_name, aliases):
    return aliases.get(normalize_team_key(raw_name), str(raw_name).strip())


def detect_format(workbook):
    for sheet_name in workbook.sheetnames:
        if sheet_name in THROWDOWN_SKIP_SHEETS:
            continue
        ws = workbook[sheet_name]
        for row in ws.iter_rows(values_only=True, max_row=30):
            if row and len(row) > 2 and row[2] and "PLAYING" in str(row[2]).upper():
                return "throwdown"
    first = workbook.worksheets[0]
    header_row = next(first.iter_rows(values_only=True, max_row=1))
    if not header_row:
        return "throwdown"
    header = [normalize_header(cell) for cell in header_row]
    if "date" in header and "court" in header and "home_team" in header:
        return "table"
    return "throwdown"


def parse_table_sheet(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_map = {}
    for idx, cell in enumerate(rows[0]):
        normalized = normalize_header(cell)
        if normalized:
            header_map[normalized] = idx

    required = {"date", "court", "home_team", "away_team", "start_time", "minutes"}
    missing = required - set(header_map)
    if missing:
        raise ValueError(
            f"Sheet '{sheet.title}' missing columns: {', '.join(sorted(missing))}. "
            f"Found headers: {list(rows[0])}"
        )

    games = []
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        game = {
            "date": normalize_date(row[header_map["date"]]),
            "court": normalize_court(row[header_map["court"]]),
            "home_team": str(row[header_map["home_team"]]).strip(),
            "away_team": str(row[header_map["away_team"]]).strip(),
            "start_time": normalize_time(row[header_map["start_time"]]),
            "minutes": int(row[header_map["minutes"]]),
        }

        if "type" in header_map and row[header_map["type"]] is not None:
            game["type"] = str(row[header_map["type"]]).strip()
        if "round" in header_map and row[header_map["round"]] is not None:
            game["round"] = str(row[header_map["round"]]).strip()

        court_num = re.search(r"(\d+)", game["court"])
        game["court_display"] = f"Court {court_num.group(1)}" if court_num else game["court"]
        games.append(game)

    return games


def parse_throwdown_workbook(workbook, date, skip_courts, minutes, include_bracket_placeholders=True, active_courts=None):
    aliases = build_team_aliases(workbook)
    round_robin_games = _parse_throwdown_team_sheet_games(workbook, date, skip_courts, minutes, aliases)
    merged, bracket_start, overview_count, bracket_count = merge_bracket_games(
        round_robin_games,
        workbook,
        date,
        skip_courts,
        minutes,
        aliases,
        resolve_team_name,
        include_placeholders=include_bracket_placeholders,
        active_courts=active_courts,
    )
    return merged, bracket_start, overview_count, bracket_count


def _parse_throwdown_team_sheet_games(workbook, date, skip_courts, minutes, aliases):
    bracket_start, bracket_time_map = build_bracket_time_map(workbook, date)
    games = []
    seen = set()

    for sheet_name in workbook.sheetnames:
        if sheet_name in THROWDOWN_SKIP_SHEETS:
            continue
        for row in workbook[sheet_name].iter_rows(values_only=True):
            if not row or len(row) < 3:
                continue
            assignment = row[2]
            if not assignment or not isinstance(assignment, str):
                continue
            upper = assignment.upper()
            if "PLAYING" not in upper:
                continue
            if "STREAM" in upper or "(AWAY)" in upper:
                continue

            match = PLAYING_HOME_RE.search(assignment)
            if not match:
                match = PLAYING_NEUTRAL_RE.search(assignment)
            if not match:
                continue

            court_num = int(match.group(1))
            if court_num in skip_courts:
                continue

            away_team = resolve_team_name(match.group(2).strip(), aliases)
            home_team = sheet_name
            round_val = row[0]
            time_val = row[1]
            if round_val is None or time_val is None:
                continue

            start_time = normalize_time(time_val)
            if not start_time:
                continue

            round_num = int(round_val)
            game_type, round_label = classify_game(
                start_time,
                round_num,
                bracket_start,
                bracket_time_map,
                bracket_minutes=minutes,
            )
            dedupe_key = (round_num, start_time, court_num, home_team, away_team)
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)

            games.append({
                "date": date,
                "court": f"court{court_num}",
                "court_display": f"Court {court_num}",
                "type": game_type,
                "round": round_label,
                "home_team": home_team,
                "away_team": away_team,
                "start_time": start_time,
                "minutes": minutes,
            })

    games.sort(key=lambda g: (g["date"], g["court"], g["start_time"], g["round"]))
    return games


def write_jsonl_files(games, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    grouped = defaultdict(list)
    for game in games:
        key = (game["date"], game["court"])
        grouped[key].append(game)

    written = []
    for (date, court), court_games in sorted(grouped.items()):
        outfile = output_dir / f"{date}_{court}.jsonl"
        with open(outfile, "w") as f:
            for game in court_games:
                record = {
                    "type": game.get("type", "round_robin"),
                    "round": game.get("round", ""),
                    "home_team": game["home_team"],
                    "away_team": game["away_team"],
                    "start_time": game["start_time"],
                    "minutes": game["minutes"],
                    "court": game["court_display"],
                }
                f.write(json.dumps(record) + "\n")
        written.append(outfile)
        print(f"Wrote {len(court_games)} games to {outfile}")

    return written


def parse_skip_courts(value):
    if not value:
        return {1}
    return {int(part.strip()) for part in value.split(",") if part.strip()}


def main():
    parser = argparse.ArgumentParser(description="Convert tournament Excel schedule to per-court games.jsonl files")
    parser.add_argument("excel_file", help="Path to tournament schedule .xlsx")
    parser.add_argument(
        "--output-dir",
        default="src/output/June2026Tournament/schedule/generated",
        help="Directory for generated JSONL files",
    )
    parser.add_argument(
        "--format",
        choices=["auto", "table", "throwdown"],
        default="auto",
        help="Excel layout (default: auto-detect)",
    )
    parser.add_argument(
        "--date",
        default="2026-06-20",
        help="Tournament date for throwdown format (default: 2026-06-20)",
    )
    parser.add_argument(
        "--minutes",
        type=int,
        default=25,
        help="Game duration in minutes for throwdown format (default: 25)",
    )
    parser.add_argument(
        "--skip-courts",
        default="1",
        help="Comma-separated courts to skip, e.g. 1 for stream court (default: 1)",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name to read for table format (default: all sheets)",
    )
    parser.add_argument(
        "--copy-to",
        default=None,
        help="Copy source Excel to this path (e.g. schedule/master_schedule.xlsx)",
    )
    parser.add_argument(
        "--no-bracket-placeholders",
        action="store_true",
        help="Skip TBD bracket slots (default: include placeholder games per court/time)",
    )
    parser.add_argument(
        "--active-courts",
        default=None,
        help="Courts with GoPro SD cards, e.g. 2,3,4 (default: date-based 2-4 Sat, 2-3 Sun)",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"Error: Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    if args.copy_to:
        copy_dest = Path(args.copy_to)
        copy_dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(excel_path, copy_dest)
        print(f"Copied schedule to {copy_dest}")

    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    fmt = args.format if args.format != "auto" else detect_format(workbook)
    skip_courts = parse_skip_courts(args.skip_courts)
    active_courts = None
    if args.active_courts:
        active_courts = [int(part.strip()) for part in args.active_courts.split(",") if part.strip()]

    if fmt == "throwdown":
        all_games, bracket_start, overview_count, bracket_count = parse_throwdown_workbook(
            workbook,
            args.date,
            skip_courts,
            args.minutes,
            include_bracket_placeholders=not args.no_bracket_placeholders,
            active_courts=active_courts,
        )
        rr_count = sum(1 for g in all_games if g["type"] == "round_robin")
        br_count = sum(1 for g in all_games if g["type"] == "bracket")
        print(
            f"Parsed {len(all_games)} games ({rr_count} round robin, {br_count} bracket) "
            f"from {bracket_start}"
        )
        if overview_count:
            print(f"  {overview_count} bracket matchup(s) from Overview Schedule")
    else:
        sheets = [workbook[args.sheet]] if args.sheet else workbook.worksheets
        all_games = []
        for sheet in sheets:
            try:
                games = parse_table_sheet(sheet)
                all_games.extend(games)
                print(f"Parsed {len(games)} games from sheet '{sheet.title}'")
            except ValueError as e:
                print(f"Skipping sheet '{sheet.title}': {e}")

    if not all_games:
        print("Error: No games parsed from Excel file.", file=sys.stderr)
        sys.exit(1)

    write_jsonl_files(all_games, args.output_dir)


if __name__ == "__main__":
    main()
